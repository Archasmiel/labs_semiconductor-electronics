from openpyxl import Workbook


def init_sheet(workbook, name):
    if name not in workbook.sheetnames:
        workbook.create_sheet(name)


wb = Workbook()
init_sheet(wb, 'test_sheet')
wb.create_sheet('test_sheet')



wb.save(filename='../!data/test.xlsx')
